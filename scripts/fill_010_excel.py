#!/usr/bin/env python3
"""Fill DS-DIN-010 excel — 9 SKU 变体 (R31-39), section 标题在 R40 (merged)."""
import openpyxl
from openpyxl.styles import Font

EXCEL = "/Users/kkkk/Library/Mobile Documents/com~apple~CloudDocs/Desktop/DSCARO.com/产品资料/01-Dining-Solutions/DS-DIN-010/DS-DIN-010_产品信息卡.xlsx"

# R33 在 010 是 SKU 变体行（不是 section 标题，R40 才是）
# 9 个 SKU 变体 R31-R39
SKU_NAMES = [
    (31, "DS-DIN-010-01", "藏青蔷薇 扣带款（下面翻折袋）", "Navy Rose Snap (Fold-up Pouch)"),
    (32, "DS-DIN-010-02", "粉色玫瑰 扣带款（下面翻折袋）", "Pink Rose Snap (Fold-up Pouch)"),
    (33, "DS-DIN-010-03", "浅灰色 扣带款（下面翻折袋）", "Light Gray Snap (Fold-up Pouch)"),
    (34, "DS-DIN-010-04", "藏蓝色 扣带款（下面翻折袋）", "Navy Blue Snap (Fold-up Pouch)"),
    (35, "DS-DIN-010-05", "蓝邹菊 扣带款（下面翻折袋）", "Blue Daisy Snap (Fold-up Pouch)"),
    (36, "DS-DIN-010-06", "蓝灰格 扣带款（下面翻折袋）", "Blue-Gray Plaid Snap (Fold-up Pouch)"),
    (37, "DS-DIN-010-07", "蓝邹菊 扣带款（上方口袋）", "Blue Daisy Snap (Top Pocket)"),
    (38, "DS-DIN-010-08", "藏青蔷薇 扣带款（上方口袋）", "Navy Rose Snap (Top Pocket)"),
    (39, "DS-DIN-010-09", "粉色玫瑰 扣带款（上方口袋）", "Pink Rose Snap (Top Pocket)"),
]

FILLS = [
    (6, 2, "Clothing Protectors（餐饮防护）"),
    (7, 2, "Snap-Button Floral Adult Bib with Pocket (扣带花卉图案围兜)"),
    (8, 2, "Health & Medical / Rehabilitation Therapy Supplies"),
    (12, 2, "https://dscaro.com/products/dining-solutions/snap-button-floral-adult-bib-din-010"),
    (15, 2, "扣带花卉图案成人围兜（9 款可选，2 种口袋设计）"),
    (16, 2, "Snap-Button Floral Adult Bib with Pocket (9 Colorways)"),
    (17, 2, "可水洗花卉图案围兜，扣带式闭合，9 种花色 + 2 种口袋（翻折袋/上方口袋），适合老年护理/失智症/居家"),
    (18, 2, "Reusable floral-pattern adult bib, snap-button closure, 9 colorways with 2 pocket designs (fold-up / top), for elderly / dementia / home care"),
    (20, 2, "标准成人码（具体尺寸建议实测）"),
    (21, 2, "花卉印花面料（floral printed fabric）+ 防水底衬（waterproof backing）"),
    (22, 2, "待确认（建议实测）"),
    (23, 2, "扣带式（Snap Button）—— 比魔术贴耐用，不夹头发"),
    (24, 2, "6 款下方翻折袋 + 3 款上方口袋（两种设计可选）"),
    (25, 2, "可机洗/手洗，可反复使用"),
    (26, 2, "面料面吸水 + 防水底"),
    (27, 2, "建议≤40°C 洗涤"),
    (28, 2, "单件 OPP 袋（建议 50 件/箱）"),
    (29, 2, "待确认"),
    (43, 2, "GBP £3.99-£4.99 (建议零售价)"),
    (48, 2, "T/T, L/C, Western Union, PayPal"),
    (49, 2, "7 天"),
    (50, 2, "15-25 天"),
    (52, 2, "Snap-Button Floral Adult Bib — 9 Colorways, Reusable Waterproof Bib with Pocket for Elderly / Dementia / Women's Care"),
    (53, 2, "snap button adult bib"),
    (54, 2, "floral elderly bib"),
    (55, 2, "reusable bib with pocket"),
    (56, 2, "dementia care bib"),
    (57, 2, "nursing home floral bib"),
    (58, 2, "待确认（建议挂 Health & Medical > Rehabilitation Therapy Supplies）"),
    (59, 2, "T/T, L/C, Western Union, PayPal"),
    (60, 2, "Dining Solutions / Elderly Care"),
    (61, 2, "15-25 Days"),
    (62, 2, "Piece / Unit"),
    (64, 2, "N/A（围兜非医疗器械）"),
    (65, 2, "待供应商提供 REACH 化学品合规声明"),
    (66, 2, "待确认（花卉印花面料可考虑 OEKO-TEX）"),
    (67, 2, "待补充（建议面料成分报告）"),
    (69, 1, "□ DS-DIN-010 定位：扣带花卉图案围兜（差异化卖点：扣带式 > 魔术贴耐用 + 9 款花纹适合女性 + 2 种口袋设计灵活）"),
    (70, 1, "□ 010 关键差异化：① 扣带式（比魔术贴耐用 3 倍，不夹头发）② 9 款花纹（女性用户友好）③ 2 种口袋（翻折袋 6 款 + 上方口袋 3 款）④ 同义乌万成 → 可与 008/009 拼单"),
    (71, 1, "□ 010 vs 009 对比：009 格子+可拆碗（$1.25-1.65，4 款）vs 010 花卉+扣带（$1.98-2.45，9 款）；010 更贵但款式更多+扣带更耐用"),
    (72, 1, "□ 010 vs 005 对比：005 双面防水（$1.98-2.65）vs 010 花卉+扣带（$1.98-2.45，便宜 8%）；010 性价比更高，005 高端"),
    (73, 1, "□ 010 适用场景：女性养老院、女性居家护理、礼品套装（多色打包）、失智症照护（色彩丰富）、医院普通餐饮"),
    (74, 1, "□ 010 包装：单件 OPP 袋 → 50件/箱；外箱唣头 DSCARO"),
    (75, 1, "□ 010 供应商：义乌万成（第 3 次合作，与 008/009 同供应商；可拼单降运费）"),
    (76, 1, "□ 010 物流：FOB Yiwu，60 件起订"),
]

GREEN = Font(color="008000")
wb = openpyxl.load_workbook(EXCEL)
ws = wb.active

# 填 9 个 SKU 变体行
sku_ok = 0
for row, sku_id, cn, en in SKU_NAMES:
    cell_a = ws.cell(row=row, column=1)
    cell_b = ws.cell(row=row, column=2)
    cell_c = ws.cell(row=row, column=3)
    if cell_a.value and "DS-DIN" in str(cell_a.value):
        cell_b.value = f"{cn} | {en}"
        cell_b.font = GREEN
        # 文件名
        sku_num = sku_id.split("-")[-1]
        cell_c.value = f"SKU_{sku_num}_{cn.split()[0]}{cn.split('（')[1].split('）')[0]}.jpg"
        cell_c.font = GREEN
        sku_ok += 1
        print(f"  {sku_id} → {cell_c.value}")

# 填其他字段
ok = 0
for row, col, val in FILLS:
    cell = ws.cell(row=row, column=col)
    if hasattr(cell, '__class__') and cell.__class__.__name__ == 'MergedCell':
        continue
    cell.value = val
    cell.font = GREEN
    ok += 1

wb.save(EXCEL)
print(f"\nSKU 变体 {sku_ok} 个 + 其他字段 {ok} 个填完")
