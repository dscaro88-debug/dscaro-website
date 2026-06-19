#!/usr/bin/env python3
"""Fill DS-DIN-012 excel — 5 SKU 变体 (R35-39), R41 是 merged section 标题."""
import openpyxl
from openpyxl.styles import Font

EXCEL = "/Users/kkkk/Library/Mobile Documents/com~apple~CloudDocs/Desktop/DSCARO.com/产品资料/01-Dining-Solutions/DS-DIN-012/DS-DIN-012_产品信息卡.xlsx"

# 5 个 SKU 变体
SKU_DATA = [
    (35, "01", "宝蓝色 / Royal Blue", "SKU_01_宝蓝色.jpg"),
    (36, "02", "灰色 / Gray", "SKU_02_灰色.jpg"),
    (37, "03", "红色 / Red", "SKU_03_红色.jpg"),
    (38, "04", "黑色 / Black", "SKU_04_黑色.jpg"),
    (39, "05", "藏蓝色 / Navy", "SKU_05_藏蓝色.jpg"),
]

FILLS = [
    (6, 2, "Clothing Protectors（餐饮防护）"),
    (7, 2, "Scarf-Style Adult Clothing Protector (围巾式成人衣物防护)"),
    (8, 2, "Health & Medical / Rehabilitation Therapy Supplies"),
    (13, 2, "https://dscaro.com/products/dining-solutions/scarf-style-adult-clothing-protector-din-012"),
    (14, 2, "本文件夹内共 36 张 jpg 图（5 主图 + 5 SKU + 26 详情），dscaro.com 选用 5 主图 + 5 SKU + 5 详情 = 15 张；svg 矢量图跳过不上传"),
    (17, 2, "围巾式成人衣物防护围兜（5 色可选）"),
    (18, 2, "Scarf-Style Adult Clothing Protector (5 Colorways)"),
    (19, 2, "围巾式套头围兜，重叠魔术贴闭合，超大尺寸 35×157cm，柔软抗污，机洗，5 色可选"),
    (20, 2, "Scarf-style drape-around adult clothing protector, overlapping hook & loop closure, extra-large 35×157cm, skin-friendly, machine washable, 5 colorways"),
    (23, 2, "35cm × 157cm（13.78in × 62in）"),
    (24, 2, "超细纤维/天鹅绒柔软面料（microfiber / velvet-soft fabric）—— 亲肤 抗污"),
    (25, 2, "待确认（建议实测）"),
    (26, 2, "重叠魔术贴闭合（Overlapping Hook & Loop）—— 无扣，套头佩戴"),
    (27, 2, "无接食袋（drape 平片式）"),
    (28, 2, "可机洗（machine washable）"),
    (29, 2, "抗污（antifouling），不吸水"),
    (30, 2, "建议≤40°C 机洗"),
    (31, 2, "单件 OPP 袋（建议 50-100 件/箱）"),
    (32, 2, "待确认"),
    (44, 2, "GBP £1.99-£2.99 (建议零售价)"),
    (49, 2, "T/T, L/C, Western Union, PayPal"),
    (50, 2, "7 天"),
    (51, 2, "15-25 天"),
    (54, 2, "Scarf-Style Adult Clothing Protector — Drape-Around Design, Overlapping Hook & Loop, Microfiber Skin-Friendly, 5 Colorways, 35×157cm Extra Large"),
    (55, 2, "scarf style adult bib"),
    (56, 2, "drape around clothing protector"),
    (57, 2, "overlapping closure bib"),
    (58, 2, "no snap elderly bib"),
    (59, 2, "microfiber adult bib"),
    (60, 2, "待确认（建议挂 Health & Medical > Rehabilitation Therapy Supplies）"),
    (61, 2, "T/T, L/C, Western Union, PayPal"),
    (62, 2, "Dining Solutions / Elderly Care"),
    (63, 2, "15-25 Days"),
    (64, 2, "Piece / Unit"),
    (67, 2, "N/A（围兜非医疗器械）"),
    (68, 2, "待供应商提供 REACH 化学品合规声明"),
    (69, 2, "待确认（microfiber 面料可考虑 OEKO-TEX）"),
    (70, 2, "待补充（建议面料成分报告）"),
    (73, 1, "□ DS-DIN-012 定位：围巾式套头围兜（差异化卖点：套头佩戴，无扣设计，适合手部不灵活者）"),
    (74, 1, "□ 012 关键差异化：① 围巾式套头 → 无扣无魔术贴繁琐 ② 重叠魔术贴 → 易穿脱 ③ 35×157cm 超大 → 全面覆盖 ④ 5 色可选 ⑤ 微纤维抗污"),
    (75, 1, "□ 012 vs 011 对比：011 加厚加大（$1.98-2.48，按扣）vs 012 围巾式（$0.98-1.28，套头）；012 便宜 50%，011 加厚耐用"),
    (76, 1, "□ 012 vs 008 对比：008 毛巾+PEVA（$1.45-1.85，魔术贴）vs 012 围巾式（$0.98-1.28，套头）；012 便宜 30%，008 吸水+防水"),
    (77, 1, "□ 012 适用场景：手部不灵活者（关节炎/中风/术后）、失智症患者、轮椅使用者、养老院（无需手部精细动作）、居家卧床护理"),
    (78, 1, "□ 012 包装：单件 OPP 袋 → 50件/箱 或 100件/箱；外箱唣头 DSCARO"),
    (79, 1, "□ 012 供应商：义乌市半夏日用百货有限公司（第 4 家新供应商，不同于 002/005 兰溪其利、006/007 金华宝舒、008/009/010/011 义乌万成）"),
    (80, 1, "□ 012 图片：根目录 36 张 jpg + 1 svg；用户已处理为英文版（AI 描述主图_04 确认是英文），无需 _en 后缀；svg 跳过"),
    (81, 1, "□ 012 链接：1688 offer 970180578376（义乌市半夏日用百货有限公司）"),
]

GREEN = Font(color="008000")
wb = openpyxl.load_workbook(EXCEL)
ws = wb.active

# 填 5 个 SKU 变体
sku_ok = 0
for row, num, cn_en, filename in SKU_DATA:
    cell_b = ws.cell(row=row, column=2)
    cell_c = ws.cell(row=row, column=3)
    cell_b.value = cn_en
    cell_b.font = GREEN
    cell_c.value = filename
    cell_c.font = GREEN
    sku_ok += 1
    print(f"  R{row} {cn_en} → {filename}")

# 填其他字段
ok = 0
for row, col, val in FILLS:
    cell = ws.cell(row=row, column=col)
    if hasattr(cell, '__class__') and cell.__class__.__name__ == 'MergedCell':
        print(f"SKIP R{row}C{col} (merged)")
        continue
    cell.value = val
    cell.font = GREEN
    ok += 1

wb.save(EXCEL)
print(f"\nSKU {sku_ok} 个 + 其他 {ok} 个填完（R41 merged 已跳过）")
